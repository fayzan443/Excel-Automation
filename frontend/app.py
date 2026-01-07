import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import io
from datetime import datetime

# ================= PAGE CONFIG =================
st.set_page_config(
    page_title="Excel Automation Tool",
    layout="wide",
    initial_sidebar_state="collapsed"
)

# ================= STYLING =================
st.markdown("""
<style>

/* REMOVE STREAMLIT AUTO HEADING LINKS */
h1 a, h2 a, h3 a, h4 a {
    display: none !important;
    pointer-events: none !important;
}

/* ---------- NAVBAR ---------- */
.navbar {
    display: flex;
    align-items: center;
    padding: 12px 50px;          /* HEIGHT SLIGHTLY INCREASED */
    margin: 15px 240px;         /* WIDTH REDUCED */
    border-radius: 60px;
    background-color: #f3f4f6;
    border-bottom: 12px solid #e5e7eb;
}
.navbar .left {
    font-weight: 600;
    font-size: 20px;
    display: flex;
    align-items: center;
    gap: 10px;
}
.navbar .right {
    display: flex;
    gap: 22px;
    margin-left: auto;
}
.navbar .right a {
    text-decoration: none;
    color: #374151;
    font-size: 15px;
}
.navbar .right a:hover {
    text-decoration: underline;
}

/* ---------- EXPANDER DEFAULT (HIDDEN) ---------- */
div[data-testid="stExpander"] {
    display: none !important;
}

/* ---------- HERO ---------- */
.hero {
    padding: 60px 40px 20px 40px;
}
.hero h1 {
    font-size: 62px;
    line-height: 1.1;
}
.hero p {
    font-size: 20px;
    max-width: 750px;
    color: #4b5563;
}

/* ---------- SECTION ---------- */
.section {
    padding: 20px 20px;
}
.section h2 {
    text-align: center;
    margin-bottom: 10px;
}
.section p {
    text-align: center;
    color: #006400;
    margin-bottom: 10px;
}

/* ---------- FEATURES ---------- */
.features {
    display: flex;
    gap: 25px;
    flex-wrap: wrap;
    justify-content: center;
}
.feature-card {
    background: #D3D3D3;
    border: 3px dotted #e5e7eb;
    border-radius: 14px;
    padding: 25px;
    width: 240px;
}
.feature-card h4,
.feature-card p {
    pointer-events: none;       /* REMOVE ANY LINK FEEL */
    text-decoration: none;
    color: #006400;
}

/* ---------- FOOTER ---------- */
.footer {
    padding: 6px;               /* DEFAULT-LIKE SMALL FOOTER */
    text-align: center;
    border-top: 2px solid #e5e7eb;
    color: #6b7280;
    font-size: 20px;
}
.footer a {
    text-decoration: none;
    color: #6b7280;
}

/* ---------- RESPONSIVE ---------- */
@media (max-width: 768px) {

    .navbar .right {
        display: none;
    }

    div[data-testid="stExpander"] {
        display: block !important;
        margin: 10px 20px 0 20px;
    }

    .navbar {
        margin: 15px 20px;
        padding: 12px 25px;
    }
}

</style>
""", unsafe_allow_html=True)

# ================= NAVBAR =================
st.markdown("""
<div class="navbar">
  <div class="left">⚙️🤖 Excel Automation</div>
  <div class="right">
    <a href="#upload">Upload</a>
    <a href="#workflow">Workflow</a>
    <a href="#results">Results</a>
    <a href="#features">Features</a>
    <a href="#how">How it works</a>
  </div>
</div>
""", unsafe_allow_html=True)

# ================= MOBILE EXPANDER =================
with st.expander("☰"):
    st.markdown("""
    <div style="text-align:center;">
        <a href="#upload">Upload</a><br>
        <a href="#workflow">Workflow</a><br>
        <a href="#results">Results</a><br>
        <a href="#features">Features</a><br>
        <a href="#how">How it works</a>
    </div>
    """, unsafe_allow_html=True)

# ================= HERO =================
st.markdown("""
<div class="hero">
  <h1>Excel Cleaner & Analyzer</h1>
  <p>
    Upload messy Excel files, clean raw data automatically, generate pivot tables,
    charts, and export professional MIS-ready reports in one click.
  </p>
</div>
""", unsafe_allow_html=True)

import streamlit as st
import pandas as pd
import numpy as np
import plotly.express as px
import io
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

# ================== HELPERS ==================

def force_numeric(series):
    series = series.astype(str).str.replace(",", "", regex=False).str.replace("%", "", regex=False)
    series = series.replace(["nan", "None", ""], np.nan)
    return pd.to_numeric(series, errors="coerce")

def clean_data(df):
    df = df.copy()
    df.columns = df.columns.str.strip().str.lower().str.replace(" ", "_")
    df = df.drop_duplicates()

    for col in df.columns:
        if df[col].dtype == "object":
            num_try = force_numeric(df[col])
            if num_try.notna().sum() > 0:
                df[col] = num_try

    for col in df.columns:
        if pd.api.types.is_numeric_dtype(df[col]):
            df[col] = df[col].fillna(df[col].median())
        else:
            df[col] = df[col].fillna("Unknown")

    return df

# ================== EXPORT FINAL EXCEL ==================
def export_final_excel(clean_df, pivot_df, chart_fig):
    output = io.BytesIO()

    # First write basic data with pandas
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        clean_df.to_excel(writer, sheet_name="Cleaned_Data", index=False)

        if pivot_df is not None:
            pivot_df.to_excel(writer, sheet_name="Pivot_Table")

        summary = pd.DataFrame({
            "Metric": ["Rows", "Columns", "Generated On"],
            "Value": [
                len(clean_df),
                len(clean_df.columns),
                datetime.now().strftime("%d-%m-%Y %H:%M")
            ]
        })
        summary.to_excel(writer, sheet_name="Summary", index=False)

    output.seek(0)
    wb = load_workbook(output)

    # ================== ADD CHART IMAGE ==================
    if chart_fig is not None:
        chart_sheet = wb.create_sheet("Charts")

        img_bytes = io.BytesIO()
        chart_fig.write_image(img_bytes, format="png", scale=2)
        img_bytes.seek(0)

        img = XLImage(img_bytes)
        chart_sheet.add_image(img, "A1")

    final = io.BytesIO()
    wb.save(final)
    final.seek(0)
    return final

# ================== SESSION STATE ==================
if 'clean_df' not in st.session_state:
    st.session_state.clean_df = None
if 'pivot_df' not in st.session_state:
    st.session_state.pivot_df = None
if 'chart_fig' not in st.session_state:
    st.session_state.chart_fig = None
if 'final_excel' not in st.session_state:
    st.session_state.final_excel = None

# ================== UPLOAD SECTION ==================
st.markdown('<h2 style="text-align: center; margin-bottom: 20px">📁 Upload Your Raw Data Excel File</h2>', unsafe_allow_html=True)

uploaded = st.file_uploader("Choose your file", type=["xlsx", "xls", "csv"], label_visibility="collapsed")

st.markdown("""
<p style="font-size:14px; color:#6b7280; margin-top:10px; text-align: center;">
• Upload ONE file at a time<br>
• Max size: 50MB<br>
• Supported formats: XLS, XLSX, CSV
</p>
""", unsafe_allow_html=True)
st.markdown('</div>', unsafe_allow_html=True)

if uploaded:
    raw_df = pd.read_excel(uploaded) if not uploaded.name.endswith(".csv") else pd.read_csv(uploaded)
    st.dataframe(raw_df.head(10))

    if st.button("🧹 Clean Data"):
        st.session_state.clean_df = clean_data(raw_df)
        st.success("Data cleaned successfully!")

# ================== ANALYSIS SECTION ==================
if st.session_state.clean_df is not None:
    df = st.session_state.clean_df
    numeric_cols = df.select_dtypes(include=[np.number]).columns.tolist()

    st.subheader("📈 Chart Builder")
    chart_type = st.selectbox("Chart Type", ["Bar", "Line", "Pie"])
    x_col = st.selectbox("X Axis", df.columns, key="x_col")
    y_col = st.selectbox("Y Axis", numeric_cols, key="y_col")

    if st.button("Generate Chart"):
        fig = px.bar(df, x=x_col, y=y_col) if chart_type == "Bar" else \
              px.line(df, x=x_col, y=y_col) if chart_type == "Line" else \
              px.pie(df, names=x_col, values=y_col)

        st.plotly_chart(fig, use_container_width=True)
        st.session_state.chart_fig = fig

    st.subheader("📊 Pivot Table")
    row = st.selectbox("Rows", df.columns, key="pivot_row")
    val = st.selectbox("Values", numeric_cols, key="pivot_val")
    agg = st.selectbox("Aggregation", ["sum", "mean", "count"], key="pivot_agg")

    if st.button("Generate Pivot Table"):
        st.session_state.pivot_df = pd.pivot_table(df, index=row, values=val, aggfunc=agg)
        st.dataframe(st.session_state.pivot_df)

# ================== FINAL REPORT ==================
st.markdown('<h2 style="margin-top:30px;">📥 Generate Final Report</h2>', unsafe_allow_html=True)

if st.button("Generate Excel Report"):
    if st.session_state.clean_df is None:
        st.warning("Please upload and clean data first!")
    else:
        st.session_state.final_excel = export_final_excel(
            st.session_state.clean_df,
            st.session_state.pivot_df,
            st.session_state.chart_fig
        )
        st.success("Final report generated!")

if st.session_state.final_excel is not None:
    st.download_button(
        "⬇ Download Final Excel Report",
        st.session_state.final_excel,
        file_name=f"Final_Report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
# ================= FEATURES =================
st.markdown('<a name="features"></a>', unsafe_allow_html=True)
st.markdown("""
<div class="section">
<h2>Key Features</h2>
<p>Designed for real office & MIS work</p>

<div class="features">
  <div class="feature-card">
    <h4>Auto Cleaning</h4>
    <p>Duplicates, missing values & formats fixed automatically</p>
  </div>
  <div class="feature-card">
    <h4>Pivot Tables</h4>
    <p>Instant summaries without Excel formulas</p>
  </div>
  <div class="feature-card">
    <h4>Charts</h4>
    <p>Bar & trend charts generated automatically</p>
  </div>
  <div class="feature-card">
    <h4>Export Ready</h4>
    <p>Download office-ready Excel reports</p>
  </div>
</div>
</div>
""", unsafe_allow_html=True)

# ================= HOW IT WORKS =================
st.markdown('<a name="how"></a>', unsafe_allow_html=True)
st.markdown("""
<div class="section">
<h2>How It Works</h2>
<p>Simple 4-step workflow</p>

<div class="features">
  <div class="feature-card" style="color: #006400;">Upload raw Excel file</div>
  <div class="feature-card" style="color: #006400;">Clean & standardize data</div>
  <div class="feature-card" style="color: #006400;">Auto-generate pivots & charts</div>
  <div class="feature-card" style="color: #006400;">Download final report</div>
</div>
</div>
""", unsafe_allow_html=True)

# ================= FOOTER =================
st.markdown("""
<div class="footer">
© 2026 Excel Automation Tool · Built for Professional Office Use  
&nbsp; | &nbsp;
<a href="mailto:muhammadfaizanofficial27@gmail.com">📧</a>
</div>
""", unsafe_allow_html=True)
